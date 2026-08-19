from django.contrib.auth.decorators import login_required
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import MovimientoInventario


def _movimientos_kardex(request):
    movimientos = MovimientoInventario.objects.select_related(
        "producto", "usuario"
    ).order_by("-fecha")
    producto_id = request.GET.get("producto")
    if producto_id:
        movimientos = movimientos.filter(producto__pk=producto_id)
    return movimientos


@login_required
def kardex(request):

    movimientos = _movimientos_kardex(request)

    contexto = {
        "movimientos": movimientos
    }


    return render(
        request,
        "reportes/kardex.html",
        contexto
    )


@login_required
def exportar_kardex_excel(request):
    movimientos = _movimientos_kardex(request)
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Kardex"
    encabezados = [
        "N°", "Fecha", "Código producto", "Producto", "Movimiento",
        "Cantidad", "Stock anterior", "Stock nuevo", "Referencia", "Usuario",
    ]
    hoja.append(encabezados)
    color = "2563EB"
    for celda in hoja[1]:
        celda.fill = PatternFill("solid", fgColor=color)
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(horizontal="center")

    for numero, movimiento in enumerate(movimientos, start=1):
        fecha = timezone.localtime(movimiento.fecha).replace(tzinfo=None)
        hoja.append([
            numero, fecha, movimiento.producto.codigo, movimiento.producto.nombre,
            movimiento.get_tipo_display(), movimiento.cantidad,
            movimiento.stock_anterior, movimiento.stock_nuevo,
            movimiento.referencia or "", movimiento.usuario.username,
        ])
        hoja.cell(row=numero + 1, column=2).number_format = "dd/mm/yyyy hh:mm"

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions
    anchos = [8, 20, 20, 34, 16, 12, 16, 14, 22, 22]
    for indice, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = ancho

    archivo = BytesIO()
    libro.save(archivo)
    archivo.seek(0)
    nombre = f"kardex_{timezone.localdate():%Y%m%d}.xlsx"
    respuesta = HttpResponse(
        archivo.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    respuesta["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return respuesta


@login_required
def api_ultimos_movimientos(request):
    """Return JSON list of recent movements. Supports optional `producto` GET param."""
    movimientos = MovimientoInventario.objects.select_related('producto', 'usuario').order_by('-fecha')[:50]
    producto_id = request.GET.get('producto')
    if producto_id:
        movimientos = movimientos.filter(producto__pk=producto_id)

    data = []
    for m in movimientos:
        data.append({
            'id': m.pk,
            'producto': m.producto.nombre,
            'producto_id': m.producto.pk,
            'tipo': m.tipo,
            'cantidad': m.cantidad,
            'fecha': m.fecha.strftime('%d/%m/%Y %H:%M'),
            'referencia': m.referencia,
            'detail_url': m.get_detail_url(),
            'stock_anterior': m.stock_anterior,
            'stock_nuevo': m.stock_nuevo,
            'usuario': getattr(m.usuario, 'username', None),
        })

    return JsonResponse({'movimientos': data})
